"""
Data Management and Reporting Features
Addresses TODO items #13-15: Data Management, Export, and Reporting
"""

import os
import csv
import json
import sqlite3
from contextlib import closing
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Any, Callable
from dataclasses import dataclass, asdict
from enum import Enum
import logging

logger = logging.getLogger(__name__)


# ============================================================================
# DATA EXPORT
# ============================================================================

class ExportFormat(Enum):
    """Supported export formats"""
    CSV = "csv"
    JSON = "json"
    EXCEL = "xlsx"


@dataclass
class ExportConfig:
    """Export configuration"""
    format: ExportFormat = ExportFormat.CSV
    include_headers: bool = True
    date_format: str = "%Y-%m-%d %H:%M:%S"
    encoding: str = "utf-8"
    delimiter: str = ","


class DataExporter:
    """
    Data export functionality
    
    Features:
    - Multiple export formats (CSV, JSON, Excel)
    - Configurable columns
    - Filtering support
    - Batch export
    """
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    def export_volunteers(
        self,
        output_path: str,
        config: Optional[ExportConfig] = None,
        filters: Optional[Dict[str, Any]] = None,
        columns: Optional[List[str]] = None
    ) -> int:
        """
        Export volunteers to file
        
        Args:
            output_path: Output file path
            config: Export configuration
            filters: Filter criteria
            columns: Columns to include
            
        Returns:
            Number of records exported
        """
        config = config or ExportConfig()
        columns = columns or [
            'profile_id', 'name', 'location', 'description',
            'skills', 'availability', 'contact_info', 'profile_url',
            'first_seen', 'last_seen', 'is_active'
        ]
        
        # Build query
        query = f"SELECT {', '.join(columns)} FROM volunteers WHERE 1=1"
        params = []
        
        if filters:
            if filters.get('location'):
                query += " AND location LIKE ?"
                params.append(f"%{filters['location']}%")
            if filters.get('active_only', True):
                query += " AND is_active = 1"
            if filters.get('since'):
                query += " AND first_seen >= ?"
                params.append(filters['since'])
        
        query += " ORDER BY last_seen DESC"
        
        # Fetch data
        with closing(sqlite3.connect(self.db_path)) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(query, params)
            rows = [dict(row) for row in cursor]
        
        # Export based on format
        if config.format == ExportFormat.CSV:
            return self._export_csv(output_path, rows, columns, config)
        elif config.format == ExportFormat.JSON:
            return self._export_json(output_path, rows, config)
        elif config.format == ExportFormat.EXCEL:
            return self._export_excel(output_path, rows, columns, config)
        
        return 0
    
    def export_messages(
        self,
        output_path: str,
        config: Optional[ExportConfig] = None,
        status: Optional[str] = None
    ) -> int:
        """Export messages to file"""
        config = config or ExportConfig()
        
        query = "SELECT * FROM scheduled_messages WHERE 1=1"
        params = []
        
        if status:
            query += " AND status = ?"
            params.append(status)
        
        query += " ORDER BY created_at DESC"
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(query, params)
            rows = [dict(row) for row in cursor]
        
        columns = [
            'id', 'recipient_id', 'recipient_name', 'subject',
            'body', 'scheduled_time', 'status', 'sent_at', 'error_message'
        ]
        
        if config.format == ExportFormat.CSV:
            return self._export_csv(output_path, rows, columns, config)
        elif config.format == ExportFormat.JSON:
            return self._export_json(output_path, rows, config)
        
        return 0
    
    def export_blacklist(
        self,
        output_path: str,
        config: Optional[ExportConfig] = None
    ) -> int:
        """Export blacklist to file"""
        config = config or ExportConfig()
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.execute("SELECT * FROM blacklist ORDER BY added_at DESC")
            rows = [dict(row) for row in cursor]
        
        columns = ['profile_id', 'name', 'reason', 'notes', 'added_at', 'is_permanent']
        
        if config.format == ExportFormat.CSV:
            return self._export_csv(output_path, rows, columns, config)
        elif config.format == ExportFormat.JSON:
            return self._export_json(output_path, rows, config)
        
        return 0
    
    def _export_csv(
        self,
        path: str,
        rows: List[Dict],
        columns: List[str],
        config: ExportConfig
    ) -> int:
        """Export to CSV"""
        with open(path, 'w', newline='', encoding=config.encoding) as f:
            writer = csv.DictWriter(
                f,
                fieldnames=columns,
                delimiter=config.delimiter,
                extrasaction='ignore'
            )
            
            if config.include_headers:
                writer.writeheader()
            
            for row in rows:
                writer.writerow(row)
        
        logger.info(f"Exported {len(rows)} records to {path}")
        return len(rows)
    
    def _export_json(
        self,
        path: str,
        rows: List[Dict],
        config: ExportConfig
    ) -> int:
        """Export to JSON"""
        with open(path, 'w', encoding=config.encoding) as f:
            json.dump(rows, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"Exported {len(rows)} records to {path}")
        return len(rows)
    
    def _export_excel(
        self,
        path: str,
        rows: List[Dict],
        columns: List[str],
        config: ExportConfig
    ) -> int:
        """Export to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Header style
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Write headers
            if config.include_headers:
                for col, header in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Write data
            start_row = 2 if config.include_headers else 1
            for row_idx, row in enumerate(rows, start_row):
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(path)
            logger.info(f"Exported {len(rows)} records to {path}")
            return len(rows)
            
        except ImportError:
            logger.error("openpyxl not installed, falling back to CSV")
            csv_path = path.replace('.xlsx', '.csv')
            return self._export_csv(csv_path, rows, columns, config)


# ============================================================================
# DATA IMPORT
# ============================================================================

class DataImporter:
    """
    Data import functionality
    
    Features:
    - Import from CSV/JSON
    - Validation
    - Duplicate handling
    - Progress tracking
    """
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    def import_volunteers(
        self,
        file_path: str,
        on_progress: Optional[Callable[[int, int], None]] = None
    ) -> Dict[str, int]:
        """
        Import volunteers from file
        
        Args:
            file_path: Input file path
            on_progress: Progress callback (current, total)
            
        Returns:
            Statistics dictionary
        """
        stats = {'imported': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
        
        # Detect format
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.csv':
            rows = self._read_csv(file_path)
        elif ext == '.json':
            rows = self._read_json(file_path)
        else:
            raise ValueError(f"Unsupported format: {ext}")
        
        total = len(rows)
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            for idx, row in enumerate(rows):
                try:
                    result = self._import_volunteer_row(conn, row)
                    stats[result] += 1
                except Exception as e:
                    logger.error(f"Error importing row {idx}: {e}")
                    stats['errors'] += 1
                
                if on_progress:
                    on_progress(idx + 1, total)
            
            conn.commit()
        
        logger.info(f"Import complete: {stats}")
        return stats
    
    def _import_volunteer_row(
        self,
        conn: sqlite3.Connection,
        row: Dict[str, Any]
    ) -> str:
        """Import a single volunteer row"""
        profile_id = row.get('profile_id', '').strip()
        
        if not profile_id:
            return 'skipped'
        
        now = datetime.now().isoformat()
        
        # Check if exists
        cursor = conn.execute(
            'SELECT id FROM volunteers WHERE profile_id = ?',
            (profile_id,)
        )
        existing = cursor.fetchone()
        
        if existing:
            # Update
            conn.execute('''
                UPDATE volunteers SET
                    name = COALESCE(?, name),
                    location = COALESCE(?, location),
                    description = COALESCE(?, description),
                    last_seen = ?,
                    last_updated = ?
                WHERE profile_id = ?
            ''', (
                row.get('name'),
                row.get('location'),
                row.get('description'),
                now,
                now,
                profile_id
            ))
            return 'updated'
        else:
            # Insert
            conn.execute('''
                INSERT INTO volunteers 
                (profile_id, name, location, description, profile_url,
                 first_seen, last_seen, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                profile_id,
                row.get('name'),
                row.get('location'),
                row.get('description'),
                row.get('profile_url'),
                now,
                now,
                now
            ))
            return 'imported'
    
    def import_blacklist(
        self,
        file_path: str,
        default_reason: str = 'other'
    ) -> Dict[str, int]:
        """Import blacklist from file"""
        stats = {'imported': 0, 'updated': 0, 'errors': 0}
        
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.csv':
            rows = self._read_csv(file_path)
        elif ext == '.json':
            rows = self._read_json(file_path)
        else:
            raise ValueError(f"Unsupported format: {ext}")
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            for row in rows:
                try:
                    profile_id = row.get('profile_id', '').strip()
                    if not profile_id:
                        continue
                    
                    now = datetime.now().isoformat()
                    
                    try:
                        conn.execute('''
                            INSERT INTO blacklist 
                            (profile_id, name, reason, notes, added_at)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (
                            profile_id,
                            row.get('name', ''),
                            row.get('reason', default_reason),
                            row.get('notes', ''),
                            now
                        ))
                        stats['imported'] += 1
                    except sqlite3.IntegrityError:
                        stats['updated'] += 1
                        
                except Exception as e:
                    logger.error(f"Error importing blacklist row: {e}")
                    stats['errors'] += 1
            
            conn.commit()
        
        return stats
    
    def _read_csv(self, path: str) -> List[Dict]:
        """Read CSV file"""
        rows = []
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
        return rows
    
    def _read_json(self, path: str) -> List[Dict]:
        """Read JSON file"""
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            elif isinstance(data, dict) and 'data' in data:
                return data['data']
        return []


# ============================================================================
# REPORTING
# ============================================================================

@dataclass
class ReportPeriod:
    """Report time period"""
    start_date: datetime
    end_date: datetime
    label: str


class ReportGenerator:
    """
    Report generation functionality
    
    Features:
    - Multiple report types
    - Customizable periods
    - Multiple output formats
    - Charts and visualizations
    """
    
    PERIODS = {
        'today': lambda: ReportPeriod(
            datetime.now().replace(hour=0, minute=0, second=0),
            datetime.now(),
            'Vandaag'
        ),
        'yesterday': lambda: ReportPeriod(
            (datetime.now() - timedelta(days=1)).replace(hour=0, minute=0, second=0),
            (datetime.now() - timedelta(days=1)).replace(hour=23, minute=59, second=59),
            'Gisteren'
        ),
        'this_week': lambda: ReportPeriod(
            datetime.now() - timedelta(days=datetime.now().weekday()),
            datetime.now(),
            'Deze Week'
        ),
        'last_week': lambda: ReportPeriod(
            datetime.now() - timedelta(days=datetime.now().weekday() + 7),
            datetime.now() - timedelta(days=datetime.now().weekday() + 1),
            'Vorige Week'
        ),
        'this_month': lambda: ReportPeriod(
            datetime.now().replace(day=1, hour=0, minute=0, second=0),
            datetime.now(),
            'Deze Maand'
        ),
        'last_month': lambda: ReportPeriod(
            (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1),
            datetime.now().replace(day=1) - timedelta(days=1),
            'Vorige Maand'
        ),
    }
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    def generate_activity_report(
        self,
        period: str = 'this_week',
        output_path: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Generate activity report
        
        Args:
            period: Report period key
            output_path: Optional output file path
            
        Returns:
            Report data dictionary
        """
        period_obj = self.PERIODS.get(period, self.PERIODS['this_week'])()
        
        report = {
            'title': 'Activiteiten Rapport',
            'period': period_obj.label,
            'generated_at': datetime.now().isoformat(),
            'start_date': period_obj.start_date.isoformat(),
            'end_date': period_obj.end_date.isoformat(),
            'sections': {}
        }
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            # Volunteer statistics
            report['sections']['volunteers'] = self._get_volunteer_stats(
                conn, period_obj
            )
            
            # Message statistics
            report['sections']['messages'] = self._get_message_stats(
                conn, period_obj
            )
            
            # Response statistics
            report['sections']['responses'] = self._get_response_stats(
                conn, period_obj
            )
            
            # Daily breakdown
            report['sections']['daily'] = self._get_daily_breakdown(
                conn, period_obj
            )
        
        # Save if output path provided
        if output_path:
            self._save_report(report, output_path)
        
        return report
    
    def _get_volunteer_stats(
        self,
        conn: sqlite3.Connection,
        period: ReportPeriod
    ) -> Dict[str, Any]:
        """Get volunteer statistics for period"""
        cursor = conn.execute('''
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN first_seen >= ? THEN 1 ELSE 0 END) as new,
                SUM(CASE WHEN last_seen >= ? THEN 1 ELSE 0 END) as active
            FROM volunteers
            WHERE is_active = 1
        ''', (period.start_date.isoformat(), period.start_date.isoformat()))
        
        row = cursor.fetchone()
        return {
            'total': row[0] or 0,
            'new_in_period': row[1] or 0,
            'active_in_period': row[2] or 0
        }
    
    def _get_message_stats(
        self,
        conn: sqlite3.Connection,
        period: ReportPeriod
    ) -> Dict[str, Any]:
        """Get message statistics for period"""
        cursor = conn.execute('''
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN status = 'sent' THEN 1 ELSE 0 END) as sent,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed,
                SUM(CASE WHEN status = 'scheduled' THEN 1 ELSE 0 END) as pending
            FROM scheduled_messages
            WHERE created_at >= ? AND created_at <= ?
        ''', (period.start_date.isoformat(), period.end_date.isoformat()))
        
        row = cursor.fetchone()
        return {
            'total': row[0] or 0,
            'sent': row[1] or 0,
            'failed': row[2] or 0,
            'pending': row[3] or 0,
            'success_rate': (row[1] / row[0] * 100) if row[0] else 0
        }
    
    def _get_response_stats(
        self,
        conn: sqlite3.Connection,
        period: ReportPeriod
    ) -> Dict[str, Any]:
        """Get response statistics for period"""
        cursor = conn.execute('''
            SELECT 
                COUNT(*) as total_reminders,
                SUM(CASE WHEN status = 'responded' THEN 1 ELSE 0 END) as responses
            FROM reminders
            WHERE created_at >= ? AND created_at <= ?
        ''', (period.start_date.isoformat(), period.end_date.isoformat()))
        
        row = cursor.fetchone()
        total = row[0] or 1
        responses = row[1] or 0
        
        return {
            'total_reminders': total,
            'responses': responses,
            'response_rate': responses / total * 100
        }
    
    def _get_daily_breakdown(
        self,
        conn: sqlite3.Connection,
        period: ReportPeriod
    ) -> List[Dict[str, Any]]:
        """Get daily breakdown for period"""
        cursor = conn.execute('''
            SELECT 
                DATE(sent_at) as date,
                COUNT(*) as messages_sent
            FROM scheduled_messages
            WHERE status = 'sent' 
            AND sent_at >= ? AND sent_at <= ?
            GROUP BY DATE(sent_at)
            ORDER BY date
        ''', (period.start_date.isoformat(), period.end_date.isoformat()))
        
        return [{'date': row[0], 'messages_sent': row[1]} for row in cursor]
    
    def _save_report(self, report: Dict, path: str) -> None:
        """Save report to file"""
        ext = os.path.splitext(path)[1].lower()
        
        if ext == '.json':
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False, default=str)
        
        elif ext == '.md':
            self._save_markdown_report(report, path)
        
        elif ext == '.html':
            self._save_html_report(report, path)
        
        logger.info(f"Report saved to {path}")
    
    def _save_markdown_report(self, report: Dict, path: str) -> None:
        """Save report as Markdown"""
        lines = [
            f"# {report['title']}",
            f"",
            f"**Periode:** {report['period']}",
            f"**Gegenereerd:** {report['generated_at']}",
            f"",
            "## Vrijwilligers",
            f"- Totaal: {report['sections']['volunteers']['total']}",
            f"- Nieuw in periode: {report['sections']['volunteers']['new_in_period']}",
            f"- Actief in periode: {report['sections']['volunteers']['active_in_period']}",
            f"",
            "## Berichten",
            f"- Totaal: {report['sections']['messages']['total']}",
            f"- Verzonden: {report['sections']['messages']['sent']}",
            f"- Mislukt: {report['sections']['messages']['failed']}",
            f"- Succes rate: {report['sections']['messages']['success_rate']:.1f}%",
            f"",
            "## Reacties",
            f"- Totaal herinneringen: {report['sections']['responses']['total_reminders']}",
            f"- Reacties: {report['sections']['responses']['responses']}",
            f"- Response rate: {report['sections']['responses']['response_rate']:.1f}%",
            f"",
            "## Dagelijks Overzicht",
            "",
            "| Datum | Berichten Verzonden |",
            "|-------|---------------------|",
        ]
        
        for day in report['sections']['daily']:
            lines.append(f"| {day['date']} | {day['messages_sent']} |")
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
    
    def _save_html_report(self, report: Dict, path: str) -> None:
        """Save report as HTML"""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>{report['title']}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; background: #1a1a2e; color: #e4e4e7; }}
        h1 {{ color: #3b82f6; }}
        h2 {{ color: #22c55e; margin-top: 30px; }}
        .stat {{ background: #1f2937; padding: 15px; margin: 10px 0; border-radius: 8px; }}
        .stat-value {{ font-size: 24px; font-weight: bold; color: #3b82f6; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #374151; padding: 10px; text-align: left; }}
        th {{ background: #0f3460; }}
    </style>
</head>
<body>
    <h1>{report['title']}</h1>
    <p><strong>Periode:</strong> {report['period']}</p>
    <p><strong>Gegenereerd:</strong> {report['generated_at']}</p>
    
    <h2>Vrijwilligers</h2>
    <div class="stat">
        <div class="stat-value">{report['sections']['volunteers']['total']}</div>
        <div>Totaal vrijwilligers</div>
    </div>
    
    <h2>Berichten</h2>
    <div class="stat">
        <div class="stat-value">{report['sections']['messages']['sent']}</div>
        <div>Verzonden ({report['sections']['messages']['success_rate']:.1f}% succes rate)</div>
    </div>
    
    <h2>Dagelijks Overzicht</h2>
    <table>
        <tr><th>Datum</th><th>Berichten Verzonden</th></tr>
        {''.join(f"<tr><td>{d['date']}</td><td>{d['messages_sent']}</td></tr>" for d in report['sections']['daily'])}
    </table>
</body>
</html>
"""
        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)


# ============================================================================
# DATA CLEANUP
# ============================================================================

class DataCleanup:
    """
    Data cleanup and maintenance
    
    Features:
    - Remove old data
    - Optimize database
    - Archive old records
    """
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    def cleanup_old_notifications(self, days: int = 30) -> int:
        """Remove notifications older than specified days"""
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.execute(
                'DELETE FROM notifications WHERE created_at < ?',
                (cutoff,)
            )
            conn.commit()
            return cursor.rowcount
    
    def cleanup_old_messages(self, days: int = 90) -> int:
        """Remove old sent messages"""
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.execute(
                'DELETE FROM scheduled_messages WHERE status = ? AND sent_at < ?',
                ('sent', cutoff)
            )
            conn.commit()
            return cursor.rowcount
    
    def mark_inactive_volunteers(self, days: int = 180) -> int:
        """Mark volunteers as inactive if not seen recently"""
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.execute(
                'UPDATE volunteers SET is_active = 0 WHERE last_seen < ? AND is_active = 1',
                (cutoff,)
            )
            conn.commit()
            return cursor.rowcount
    
    def optimize_database(self) -> None:
        """Optimize database (vacuum and analyze)"""
        with closing(sqlite3.connect(self.db_path)) as conn:
            conn.execute('VACUUM')
            conn.execute('ANALYZE')
        
        logger.info("Database optimized")
    
    def get_database_stats(self) -> Dict[str, Any]:
        """Get database statistics"""
        stats = {}
        
        # File size
        if os.path.exists(self.db_path):
            stats['file_size_mb'] = os.path.getsize(self.db_path) / (1024 * 1024)
        
        with closing(sqlite3.connect(self.db_path)) as conn:
            # Table counts
            tables = ['volunteers', 'scheduled_messages', 'reminders', 
                     'blacklist', 'notifications']
            
            for table in tables:
                try:
                    cursor = conn.execute(f'SELECT COUNT(*) FROM {table}')
                    stats[f'{table}_count'] = cursor.fetchone()[0]
                except:
                    stats[f'{table}_count'] = 0
        
        return stats

